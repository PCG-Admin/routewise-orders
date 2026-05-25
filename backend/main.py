from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from supabase import create_client, Client
import os
from dotenv import load_dotenv
from datetime import datetime
import uuid
import io
import json
import re
import logging

from utils import safe_date, safe_float
from pdf_extractor import extract_pdf_text
from excel_extractor import extract_excel_structured, excel_to_text

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

load_dotenv()

app = FastAPI(title="Bulk Connections API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:3001",
        "http://localhost:3002",
        "https://bulk-01.vercel.app",
        "https://bulk-01-git-main-alisonrajpals-projects.vercel.app",
        "https://bulk-01-etlzmubpx-alisonrajpals-projects.vercel.app",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

try:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    print("Supabase connected!")
except Exception as e:
    print(f"Supabase error: {e}")
    supabase = None

# Gemini client (new google.genai SDK)
gemini_client = None
if GEMINI_API_KEY:
    try:
        from google import genai as _genai  # type: ignore[import]
        gemini_client = _genai.Client(api_key=GEMINI_API_KEY)
        print("Gemini initialized!")
    except Exception as e:
        print(f"Gemini error: {e}")

# ============================================
# ROOT & HEALTH
# ============================================

@app.get("/")
async def root():
    return {"message": "API is running", "status": "healthy"}

@app.get("/api/health")
async def health():
    return {"status": "ok", "timestamp": datetime.now().isoformat()}

# ============================================
# AUTHENTICATION
# ============================================

@app.post("/api/auth/login")
async def login(credentials: dict):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase not connected")

    email = (credentials.get("email") or "").strip().lower()
    password = credentials.get("password") or ""

    if not email or not password:
        raise HTTPException(status_code=400, detail="Email and password are required")

    try:
        result = supabase.table("users").select("id,email,name,role,password").eq("email", email).execute()
        if not result.data:
            raise HTTPException(status_code=401, detail="Invalid email or password")

        user = result.data[0]
        if user.get("password") != password:
            raise HTTPException(status_code=401, detail="Invalid email or password")

        return {
            "id": user["id"],
            "email": user["email"],
            "name": user.get("name") or "",
            "role": user.get("role") or "user",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Login error: {e}")
        raise HTTPException(status_code=500, detail="Login failed")

# ============================================
# ORDERS CRUD
# ============================================

@app.get("/api/orders")
async def get_orders():
    if not supabase:
        return {"orders": [], "error": "Supabase not connected"}
    try:
        result = supabase.table("orders").select("*").order("createdAt", desc=True).execute()
        return {"orders": result.data}
    except Exception as e:
        return {"orders": [], "error": str(e)}

@app.get("/api/orders/{order_id}")
async def get_order(order_id: str):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase not connected")
    try:
        result = supabase.table("orders").select("*").eq("id", order_id).execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="Order not found")
        return result.data[0]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/orders")
async def create_order(order_data: dict):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase not connected")
    try:
        new_order = {
            "id": str(uuid.uuid4()),
            "orderNumber": order_data.get("orderNumber", f"ORD-{int(datetime.now().timestamp())}"),
            "clientName": order_data.get("clientName", ""),
            "product": order_data.get("product", ""),
            "quantity": float(order_data.get("quantity", 0)) if order_data.get("quantity") else 0,
            "unit": order_data.get("unit", "kg"),
            "status": order_data.get("status", "pending"),
            "priority": order_data.get("priority", "normal"),
            "originAddress": order_data.get("originAddress", ""),
            "destinationAddress": order_data.get("destinationAddress", ""),
            "notes": order_data.get("notes", ""),
            "requestedPickupDate": order_data.get("requestedPickupDate"),
            "requestedDeliveryDate": order_data.get("requestedDeliveryDate"),
            "createdAt": datetime.now().isoformat(),
            "updatedAt": datetime.now().isoformat(),
        }
        new_order = {k: v for k, v in new_order.items() if v is not None}
        result = supabase.table("orders").insert(new_order).execute()
        return {"message": "Order created successfully", "order": result.data[0] if result.data else new_order}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/orders/{order_id}")
async def update_order(order_id: str, order_data: dict):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase not connected")
    try:
        update_data = {k: v for k, v in order_data.items() if v is not None}
        result = supabase.table("orders").update(update_data).eq("id", order_id).execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="Order not found")
        return {"message": "Order updated successfully", "order": result.data[0]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/orders/{order_id}")
async def delete_order(order_id: str):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase not connected")
    try:
        supabase.table("truck_allocations").delete().eq("orderId", order_id).execute()
        result = supabase.table("orders").delete().eq("id", order_id).execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="Order not found")
        return {"message": "Order deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# TRUCK ALLOCATIONS CRUD
# ============================================

@app.get("/api/orders/{order_id}/trucks")
async def get_order_trucks(order_id: str):
    if not supabase:
        return {"trucks": []}
    try:
        result = supabase.table("truck_allocations").select("*").eq("orderId", order_id).execute()
        trucks = []
        for item in result.data:
            trucks.append({
                "id": item.get("id"),
                "plate": item.get("vehicleReg", ""),
                "status": item.get("status", "scheduled"),
                "driver": item.get("driverName", ""),
                "phone": item.get("driverPhone", ""),
                "transporter": item.get("transporter", ""),
                "scheduledDate": item.get("scheduledDate", ""),
                "ticketNo": item.get("ticketNo", ""),
                "netWeight": item.get("netWeight", 0),
                "grossWeight": item.get("grossWeight", 0),
                "tareWeight": item.get("tareWeight", 0),
                "fleetNo": item.get("fleetNo", ""),
                "trailer1": item.get("trailer1", ""),
                "trailer2": item.get("trailer2", ""),
            })
        return {"trucks": trucks}
    except Exception as e:
        logger.error(f"Error fetching trucks: {e}")
        return {"trucks": []}

@app.post("/api/orders/{order_id}/trucks")
async def add_truck_to_order(order_id: str, truck_data: dict):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase not connected")
    try:
        new_truck = {
            "id": str(uuid.uuid4()),
            "orderId": order_id,
            "vehicleReg": truck_data.get("plate", ""),
            "driverName": truck_data.get("driver", ""),
            "driverPhone": truck_data.get("phone", ""),
            "transporter": truck_data.get("transporter", ""),
            "status": truck_data.get("status", "scheduled"),
            "scheduledDate": truck_data.get("scheduledDate"),
            "ticketNo": truck_data.get("ticketNo", ""),
            "netWeight": float(truck_data.get("netWeight", 0)) if truck_data.get("netWeight") else None,
        }
        new_truck = {k: v for k, v in new_truck.items() if v is not None and v != ""}
        result = supabase.table("truck_allocations").insert(new_truck).execute()
        return {"message": "Truck added successfully", "truck": result.data[0] if result.data else new_truck}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/orders/{order_id}/trucks/{truck_id}")
async def update_truck_allocation(order_id: str, truck_id: str, truck_data: dict):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase not connected")
    try:
        update_data = {k: v for k, v in truck_data.items() if v is not None}
        supabase.table("truck_allocations").update(update_data).eq("id", truck_id).eq("orderId", order_id).execute()
        return {"message": "Truck updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/orders/{order_id}/trucks/{truck_id}")
async def delete_truck_allocation(order_id: str, truck_id: str):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase not connected")
    try:
        supabase.table("truck_allocations").delete().eq("id", truck_id).eq("orderId", order_id).execute()
        return {"message": "Truck deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# SHARED HELPERS
# ============================================

def save_extracted_to_db(extracted: dict, filename: str) -> dict:
    """Save Gemini-extracted order + trucks to Supabase. Returns summary."""
    order_number = (extracted.get("orderNumber") or "").strip() or f"ORD-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    trucks_list = extracted.get("trucks", [])

    # Check if order already exists
    existing = supabase.table("orders").select("id,orderNumber").eq("orderNumber", order_number).execute()
    if existing.data:
        order = existing.data[0]
        logger.info(f"Order already exists: {order_number}")
    else:
        new_order = {
            "id": str(uuid.uuid4()),
            "orderNumber": order_number,
            "clientName": extracted.get("clientName") or "Unknown Client",
            "product": extracted.get("product") or "Unknown",
            "quantity": safe_float(extracted.get("orderQty")) or float(len(trucks_list)),
            "unit": extracted.get("unit") or "tons",
            "status": "pending",
            "priority": "normal",
            "originAddress": extracted.get("originAddress") or "",
            "destinationAddress": extracted.get("destinationAddress") or "",
            "requestedPickupDate": safe_date(extracted.get("pickupDate")),
            "notes": f"Imported from {filename}. Trucks: {len(trucks_list)}",
            "createdAt": datetime.now().isoformat(),
            "updatedAt": datetime.now().isoformat(),
        }
        new_order = {k: v for k, v in new_order.items() if v is not None and v != ""}
        result = supabase.table("orders").insert(new_order).execute()
        if not result.data:
            raise Exception("Failed to create order in database")
        order = result.data[0]
        logger.info(f"Created order: {order_number} id={order['id']}")

    # Insert truck allocations
    trucks_added = 0
    trucks_skipped = 0
    for t in trucks_list:
        vehicle_reg = (t.get("vehicleReg") or "").strip().upper()
        if not vehicle_reg:
            trucks_skipped += 1
            continue
        try:
            truck_row = {
                "id": str(uuid.uuid4()),
                "orderId": order["id"],
                "vehicleReg": vehicle_reg,
                "status": t.get("status") or "scheduled",
            }
            if t.get("driverName"):
                truck_row["driverName"] = str(t["driverName"]).strip()
            if t.get("transporter"):
                truck_row["transporter"] = str(t["transporter"]).strip()
            if t.get("fleetNo"):
                truck_row["fleetNo"] = str(t["fleetNo"]).strip()
            if t.get("trailer1"):
                truck_row["trailer1"] = str(t["trailer1"]).strip().upper()
            if t.get("trailer2"):
                truck_row["trailer2"] = str(t["trailer2"]).strip().upper()
            if t.get("driverId"):
                truck_row["driverId"] = str(t["driverId"]).strip()
            if t.get("ticketNo"):
                truck_row["ticketNo"] = str(t["ticketNo"]).strip()
            if safe_float(t.get("grossWeight")) is not None:
                truck_row["grossWeight"] = safe_float(t["grossWeight"])
            if safe_float(t.get("tareWeight")) is not None:
                truck_row["tareWeight"] = safe_float(t["tareWeight"])
            if safe_float(t.get("netWeight")) is not None:
                truck_row["netWeight"] = safe_float(t["netWeight"])
            if safe_date(t.get("scheduledDate")):
                truck_row["scheduledDate"] = safe_date(t["scheduledDate"])

            supabase.table("truck_allocations").insert(truck_row).execute()
            trucks_added += 1
        except Exception as e:
            logger.error(f"Failed to insert truck {vehicle_reg}: {e}")
            trucks_skipped += 1

    return {
        "orderId": order["id"],
        "orderNumber": order_number,
        "trucks_added": trucks_added,
        "trucks_skipped": trucks_skipped,
    }

# ============================================
# GEMINI AI EXTRACTION
# ============================================

GEMINI_PROMPT = """You are a data extraction AI for a South African bulk transport management system.

Parse the document and return ONLY a valid JSON object (no markdown, no code fences, just raw JSON):

{
  "orderNumber": "order number string or null",
  "clientName": "mine or company name or null",
  "product": "product code e.g. CHR001 or null",
  "orderQty": total_quantity_tonnes_as_number_or_null,
  "unit": "tons",
  "originAddress": "source mine/site or null",
  "destinationAddress": "destination port or city or null",
  "pickupDate": "YYYY-MM-DD or null",
  "trucks": [
    {
      "vehicleReg": "truck registration plate - REQUIRED",
      "driverName": "driver full name or null",
      "transporter": "transport company name or null",
      "fleetNo": "fleet number or null",
      "trailer1": "first trailer plate or null",
      "trailer2": "second trailer plate or null",
      "driverId": "driver ID number or null",
      "ticketNo": "weighbridge ticket or transaction number or null",
      "grossWeight": gross_kg_as_integer_or_null,
      "tareWeight": tare_kg_as_integer_or_null,
      "netWeight": net_kg_as_integer_or_null,
      "scheduledDate": "YYYY-MM-DD or null",
      "status": "scheduled"
    }
  ]
}

WEIGHBRIDGE DETAIL REPORT (WDR) — Newton system used at South African mines:
- ORDER NUMBER: extract from the "ORDER NUMBER:" summary box at the bottom of the report (e.g. ELD26-05-30, ZDE26-05-46), NOT from the ORDERNO column in the table rows
- clientName: from the report title header e.g. "NORTHAM ELAND", "FARM ZONDEREINDE"
- originAddress: mine name from the report title
- orderQty: from "ORDER QTY:" in the summary box (value is already in tonnes)
- pickupDate: date portion of the "Filter Range" start date, format YYYY-MM-DD
- Each DISP transaction row = one truck entry:
  - TRUCK REG column → vehicleReg (e.g. LGZ388MP)
  - TRAN NO column → ticketNo (e.g. 1001301 or 136262)
  - TARE column → tareWeight in kg (e.g. 19050)
  - GROSS column → grossWeight in kg (e.g. 56950)
  - NETT column → netWeight in kg (e.g. 37900)
  - TRANSPORTER column → transporter (e.g. VRC)
  - PRODUCT column → product (e.g. CHR001)
  - DEST column → destinationAddress (e.g. DBN, BC)
  - SUPPLIER column → originAddress on truck row (can leave null, use report-level originAddress)
  - USER column is a weighbridge OPERATOR, NOT the truck driver — set driverName to null
  - Set status = "completed" for all WDR trucks (they have been weighed and dispatched)

GENERAL RULES:
- Include ALL transaction rows — do not skip or summarise any
- vehicleReg is REQUIRED — skip rows with no truck registration
- Weight values must be plain integers. No units
- All dates must be YYYY-MM-DD
- Return ONLY raw JSON — no explanation, no markdown

DOCUMENT:
"""

def _parse_gemini_response(raw_text: str) -> dict | None:
    """Strip markdown fences and parse JSON from a Gemini response."""
    try:
        raw = raw_text.strip()
        raw = re.sub(r'^```(?:json)?\s*', '', raw, flags=re.IGNORECASE)
        raw = re.sub(r'\s*```$', '', raw)
        data = json.loads(raw.strip())
        logger.info(f"Gemini parsed: {len(data.get('trucks', []))} trucks, order={data.get('orderNumber')}")
        return data
    except json.JSONDecodeError as e:
        logger.error(f"Gemini JSON parse error: {e} — raw: {raw_text[:400]}")
        return None

def call_gemini(text: str) -> dict | None:
    """Call Gemini with plain text. Returns parsed dict or None."""
    if not gemini_client:
        return None
    if not text or len(text.strip()) < 50:
        return None
    try:
        response = gemini_client.models.generate_content(
            model="gemini-2.5-flash",
            contents=GEMINI_PROMPT + text[:30000]
        )
        return _parse_gemini_response(response.text)
    except Exception as e:
        logger.error(f"Gemini text call failed: {e}")
        return None


# ============================================
# PDF UPLOAD — PyPDF2 → OCR FALLBACK → GEMINI 2.5 FLASH
# ============================================

@app.post("/api/upload/pdf")
async def upload_pdf(file: UploadFile = File(...)):
    """
    Upload a weighbridge / WDR PDF.
    Pipeline:
      1. PyPDF2  — fast text extraction for digital PDFs
      2. OCR     — pdf2image + pytesseract when PyPDF2 yields < 200 chars
      3. Gemini 2.5 Flash — structured data extraction from the text
      4. Supabase — save order + truck allocations
    """
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase not connected")

    logger.info(f"PDF upload started: {file.filename}")
    try:
        contents = await file.read()

        # ── Steps 1 & 2: PyPDF2 → OCR fallback (handled in pdf_extractor) ─
        pdf_text = extract_pdf_text(contents)

        if not pdf_text.strip():
            raise HTTPException(
                status_code=400,
                detail=(
                    "Could not extract text from this PDF. "
                    "For scanned PDFs, ensure Tesseract OCR is installed on the server."
                ),
            )

        logger.info(f"Text ready for Gemini: {len(pdf_text)} chars")

        # ── Step 3: Gemini 2.5 Flash structured extraction ────────────────
        extracted = call_gemini(pdf_text)
        if not extracted:
            raise HTTPException(
                status_code=422,
                detail=(
                    "Gemini could not parse this PDF. "
                    "Ensure it is a Newton weighbridge report (WDR) or similar order document."
                ),
            )

        trucks_found = len(extracted.get("trucks", []))
        logger.info(
            f"Gemini extracted: order={extracted.get('orderNumber')}, trucks={trucks_found}"
        )

        # ── Step 4: Save to Supabase ───────────────────────────────────────
        summary = save_extracted_to_db(extracted, file.filename)
        return {
            "success": True,
            "message": f"Extracted {summary['trucks_added']} trucks for order {summary['orderNumber']}",
            **summary,
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# STATS & REPORTING
# ============================================

@app.get("/api/stats")
async def get_stats():
    """Aggregated stats for all report pages — orders + truck allocations."""
    if not supabase:
        return {}
    try:
        orders = supabase.table("orders").select("*").execute().data or []
        trucks = supabase.table("truck_allocations").select("*").execute().data or []

        total_orders = len(orders)
        active_orders = sum(1 for o in orders if o.get("status") not in ("completed", "cancelled"))
        total_trucks = len(trucks)

        # Truck status breakdown
        truck_status: dict = {}
        for t in trucks:
            s = t.get("status") or "scheduled"
            truck_status[s] = truck_status.get(s, 0) + 1

        # Product mix from orders
        product_counts: dict = {}
        for o in orders:
            p = (o.get("product") or "Unknown").strip()
            product_counts[p] = product_counts.get(p, 0) + 1
        product_mix = [{"name": k, "value": v} for k, v in sorted(product_counts.items(), key=lambda x: -x[1])]

        # Completed/total trucks per order
        order_truck_map: dict = {}
        for t in trucks:
            oid = t.get("orderId")
            if not oid:
                continue
            if oid not in order_truck_map:
                order_truck_map[oid] = {"total": 0, "completed": 0}
            order_truck_map[oid]["total"] += 1
            if t.get("status") == "completed":
                order_truck_map[oid]["completed"] += 1

        order_completions = [
            {
                "id": o.get("orderNumber") or o["id"][:8],
                "orderNumber": o.get("orderNumber"),
                "clientName": o.get("clientName"),
                "completed": order_truck_map.get(o["id"], {}).get("completed", 0),
                "total": order_truck_map.get(o["id"], {}).get("total", 0),
            }
            for o in orders
        ]

        # Transporter stats
        transporter_map: dict = {}
        for t in trucks:
            name = (t.get("transporter") or "Unknown").strip()
            if name not in transporter_map:
                transporter_map[name] = {"total": 0, "completed": 0}
            transporter_map[name]["total"] += 1
            if t.get("status") == "completed":
                transporter_map[name]["completed"] += 1
        transporter_stats = sorted(
            [
                {
                    "name": k,
                    "total": v["total"],
                    "completed": v["completed"],
                    "score": round((v["completed"] / v["total"]) * 100) if v["total"] > 0 else 0,
                }
                for k, v in transporter_map.items()
            ],
            key=lambda x: -x["total"],
        )

        # Client/mine stats
        client_map: dict = {}
        for o in orders:
            c = (o.get("clientName") or "Unknown").strip()
            if c not in client_map:
                client_map[c] = {"orders": 0, "trucks": 0, "tonnes": 0.0, "product": o.get("product") or ""}
            client_map[c]["orders"] += 1
            client_map[c]["trucks"] += order_truck_map.get(o["id"], {}).get("total", 0)
            client_map[c]["tonnes"] += float(o.get("quantity") or 0)
        client_stats = sorted(
            [{"name": k, "orders": v["orders"], "trucks": v["trucks"], "tonnes": round(v["tonnes"], 2), "product": v["product"]}
             for k, v in client_map.items()],
            key=lambda x: -x["tonnes"],
        )

        return {
            "totalOrders": total_orders,
            "activeOrders": active_orders,
            "totalTrucks": total_trucks,
            "truckStatus": truck_status,
            "productMix": product_mix,
            "orderCompletions": order_completions,
            "transporterStats": transporter_stats,
            "clientStats": client_stats,
        }
    except Exception as e:
        logger.error(f"Stats error: {e}")
        return {}


@app.get("/api/trucks")
async def get_all_trucks():
    """All truck allocations across all orders (latest 500)."""
    if not supabase:
        return {"trucks": []}
    try:
        result = supabase.table("truck_allocations").select("*").limit(500).execute()
        return {"trucks": result.data or []}
    except Exception as e:
        logger.error(f"Error fetching all trucks: {e}")
        return {"trucks": []}

# ============================================
# EXCEL UPLOAD — TEMPLATE ENGINE + GEMINI AI
# ============================================

@app.post("/api/upload/excel")
async def upload_excel(file: UploadFile = File(...)):
    """Upload Excel file — template extraction with Gemini AI fallback."""
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase not connected")

    logger.info(f"Excel upload: {file.filename}")
    try:
        contents = await file.read()

        # Step 1: Try structured template extraction (openpyxl)
        extracted = extract_excel_structured(contents)

        # Step 2: Fall back to Gemini AI if no template matched
        if not extracted:
            excel_text = excel_to_text(contents)
            if not excel_text.strip():
                raise HTTPException(status_code=400, detail="Could not read Excel file")
            logger.info(f"Gemini fallback — Excel text ({len(excel_text)} chars):\n{excel_text[:500]}")
            extracted = call_gemini(excel_text)

        if not extracted:
            raise HTTPException(
                status_code=422,
                detail="Could not parse Excel. Ensure the file matches a supported format or use the provided template.",
            )

        # Step 3: Save to database
        summary = save_extracted_to_db(extracted, file.filename)

        return {
            "success": True,
            "message": f"Extracted {summary['trucks_added']} trucks for order {summary['orderNumber']}",
            **summary,
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# EXCEL TEMPLATE DOWNLOAD
# ============================================

@app.get("/api/template/excel")
async def download_excel_template():
    """Download the standard truck allocation Excel template."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Truck Allocation"

        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        info_fill = PatternFill(start_color="EBF3FF", end_color="EBF3FF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        info_font = Font(bold=True, color="1E3A5F", size=10)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")

        # Row 1: Title
        ws.merge_cells("A1:H1")
        ws["A1"] = "BULK CONNECTIONS — TRUCK ALLOCATION SHEET"
        ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 30

        # Rows 3-8: Order info fields
        order_fields = [
            ("Order Number:", "e.g. VRC-2026-05-001"),
            ("Client / Mine Name:", "e.g. Northam Eland"),
            ("Product:", "e.g. Chrome Ore / CHR001"),
            ("Total Quantity (tons):", "e.g. 5000"),
            ("Origin (Mine/Site):", "e.g. Northam Eland Mine"),
            ("Destination:", "e.g. Richards Bay Port"),
            ("Pickup Date:", "e.g. 2026-05-13"),
        ]
        for i, (label, placeholder) in enumerate(order_fields, start=3):
            ws.cell(row=i, column=1, value=label).font = info_font
            ws.cell(row=i, column=1).fill = info_fill
            ws.cell(row=i, column=1).alignment = Alignment(horizontal="right", vertical="center")
            cell = ws.cell(row=i, column=2, value=placeholder)
            cell.font = Font(color="999999", italic=True, size=10)
            ws.merge_cells(f"B{i}:D{i}")

        ws.row_dimensions[10].height = 8  # spacer

        # Row 11: Column headers for truck data
        truck_headers = [
            "TRANSPORTER", "FLEET NO", "HORSE REG *", "TRAILER 1",
            "TRAILER 2", "DRIVER NAME", "DRIVER ID/NO", "SCHEDULED DATE"
        ]
        for col, header in enumerate(truck_headers, start=1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[11].height = 22

        # Sample data rows
        sample_rows = [
            ["VRC Transport", "F001", "LGZ388MP", "TR001", "TR002", "John Dlamini", "8901015800089", "2026-05-13"],
            ["VRC Transport", "F002", "LHT407MP", "TR003", "", "Peter Nkosi", "7805125800082", "2026-05-13"],
            ["VRC Transport", "F003", "LFC145MP", "", "", "Mary Sithole", "9002034800081", "2026-05-14"],
        ]
        sample_fill = PatternFill(start_color="F8FBFF", end_color="F8FBFF", fill_type="solid")
        for row_idx, row_data in enumerate(sample_rows, start=12):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = sample_fill
                cell.border = border
                cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[row_idx].height = 18

        # Add 20 blank data rows
        blank_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row_idx in range(15, 35):
            for col_idx in range(1, 9):
                ws.cell(row=row_idx, column=col_idx).fill = blank_fill
                ws.cell(row=row_idx, column=col_idx).border = border
            ws.row_dimensions[row_idx].height = 18

        # Column widths
        col_widths = [22, 12, 14, 12, 12, 20, 18, 16]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Note row at bottom
        note_row = 36
        ws.merge_cells(f"A{note_row}:H{note_row}")
        ws.cell(row=note_row, column=1, value="* HORSE REG is required for each row. Delete sample rows before uploading.").font = Font(color="FF0000", italic=True, size=9)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=truck_allocation_template.xlsx"}
        )
    except Exception as e:
        logger.error(f"Template generation error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# EXCEL REPORT — REAL DATA FROM DATABASE
# ============================================

@app.get("/api/report/excel/{order_id}")
async def download_order_report(order_id: str):
    """Generate a filled KFTS-style Excel report for an order using live DB data."""
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase not connected")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        order_res = supabase.table("orders").select("*").eq("id", order_id).execute()
        if not order_res.data:
            raise HTTPException(status_code=404, detail="Order not found")
        order = order_res.data[0]

        trucks_res = supabase.table("truck_allocations").select("*").eq("orderId", order_id).execute()
        trucks = trucks_res.data or []

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        order_number = order.get("orderNumber", "")
        destination = order.get("destinationAddress") or order.get("clientName") or ""
        pickup_date = order.get("requestedPickupDate", "") or ""

        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alt_fill = PatternFill(start_color="F0F4FA", end_color="F0F4FA", fill_type="solid")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")

        # Row 1: KFTS-style metadata (job ref left, loaded date right)
        row1_left = f"{order_number} – {destination}".strip(" –")
        row1_right = f"LOADED {pickup_date}" if pickup_date else "LOADED —"

        ws.merge_cells("A1:D1")
        ws["A1"] = row1_left
        ws["A1"].font = Font(bold=True, size=12, color="1E3A5F")

        ws.merge_cells("E1:G1")
        ws["E1"] = row1_right
        ws["E1"].font = Font(bold=True, size=12, color="1E3A5F")
        ws["E1"].alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[1].height = 24

        # Row 3: column headers
        col_headers = ["TRANSPORTER", "FLEET NO", "HORSE REG", "TRAILER 1", "TRAILER 2", "DRIVER NAME SURNAME", "ID"]
        for col, h in enumerate(col_headers, start=1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[3].height = 22

        # Rows 4+: truck data from DB
        for row_idx, truck in enumerate(trucks, start=4):
            row_data = [
                truck.get("transporter", ""),
                truck.get("fleetNo", ""),
                truck.get("vehicleReg", ""),
                truck.get("trailer1", ""),
                truck.get("trailer2", ""),
                truck.get("driverName", ""),
                truck.get("driverId", ""),
            ]
            fill = alt_fill if row_idx % 2 == 0 else None
            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                if fill:
                    cell.fill = fill
            ws.row_dimensions[row_idx].height = 18

        for col, width in enumerate([22, 12, 14, 12, 12, 26, 18], start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_num = re.sub(r'[^\w\-]', '_', order_number) or order_id[:8]
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={safe_num}_report.xlsx"},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Report generation error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
