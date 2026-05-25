import io
import re
import logging
import pandas as pd
from utils import safe_date, safe_float

logger = logging.getLogger(__name__)


# ── Low-level helpers ────────────────────────────────────────────────────────

def _first_row_texts(ws, row: int) -> list[str]:
    """Return all distinct non-empty string values from a worksheet row (skips merged-cell duplicates)."""
    seen, results = set(), []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            t = str(v).strip()
            if t and t not in seen:
                seen.add(t)
                results.append(t)
    return results


def _map_headers(header_list: list[str]) -> dict[str, int]:
    """Map normalised header strings → 0-based column index."""
    lookup = {
        "TRANSPORTER": "transporter",
        "FLEET NO": "fleetNo", "FLEET": "fleetNo",
        "HORSE REG": "vehicleReg", "HORSE": "vehicleReg", "VEHICLE REG": "vehicleReg",
        "TRAILER 1": "trailer1", "TRAILER1": "trailer1",
        "TRAILER 2": "trailer2", "TRAILER2": "trailer2",
        "DRIVER NAME SURNAME": "driverName", "DRIVER NAME": "driverName", "DRIVER": "driverName",
        "ID": "driverId", "DRIVER ID": "driverId", "DRIVER ID/NO": "driverId",
        "SCHEDULED DATE": "scheduledDate", "DATE": "scheduledDate",
        "TICKET NO": "ticketNo",
        "GROSS WEIGHT": "grossWeight", "GROSS": "grossWeight",
        "TARE WEIGHT": "tareWeight",  "TARE": "tareWeight",
        "NET WEIGHT": "netWeight",    "NET": "netWeight",
    }
    mapping = {}
    for idx, h in enumerate(header_list):
        key = str(h).strip().upper().rstrip(" *:")
        if key in lookup:
            mapping[lookup[key]] = idx
    return mapping


def _read_truck_row(ws, row: int, col_map: dict, num_cols: int) -> dict | None:
    """Read one data row; return a truck dict or None if the row is fully empty."""
    row_vals = [ws.cell(row=row, column=c).value for c in range(1, num_cols + 1)]
    if all(v is None or str(v).strip() == "" for v in row_vals):
        return None
    truck: dict = {"status": "scheduled"}
    for field, idx in col_map.items():
        v = row_vals[idx] if idx < len(row_vals) else None
        if v is None or str(v).strip() == "":
            continue
        if field in ("grossWeight", "tareWeight", "netWeight"):
            truck[field] = safe_float(v)
        elif field == "fleetNo":
            s = str(v).strip()
            truck[field] = str(int(float(s))) if re.fullmatch(r'\d+(\.\d+)?', s) else s
        else:
            truck[field] = str(v).strip()
    return truck if truck.get("vehicleReg") else None


# ── Template 1: KFTS / Island View ──────────────────────────────────────────

def _extract_kfts_template(ws) -> dict | None:
    """
    Row 1 — merged metadata: job reference + 'LOADED DD-MM-YYYY'
    Row 2 — empty
    Row 3 — column headers (TRANSPORTER, FLEET NO, HORSE REG, …)
    Row 4+ — one truck per row; stops at first fully-empty row
    """
    row1_texts = _first_row_texts(ws, 1)
    job_ref, destination, loaded_date = "", "", None
    non_loaded: list[str] = []

    for text in row1_texts:
        m_date = re.search(r'LOADED\s+(\d{1,2}[-/]\d{1,2}[-/]\d{4})', text, re.IGNORECASE)
        if m_date:
            loaded_date = safe_date(m_date.group(1))
            remainder = re.sub(r'LOADED\s+\d{1,2}[-/]\d{1,2}[-/]\d{4}', '', text, flags=re.IGNORECASE).strip()
            if remainder:
                non_loaded.append(remainder)
        else:
            non_loaded.append(text)

    if non_loaded:
        parts = re.split(r'\s*[–—]\s*|\s+-\s+', non_loaded[0], maxsplit=1)
        if len(parts) > 1:
            job_ref = parts[0].strip()
            destination = parts[1].strip()
        else:
            m_ref = re.match(r'^(\S+)\s+(.+)', non_loaded[0])
            if m_ref and re.search(r'\d', m_ref.group(1)):
                job_ref = m_ref.group(1).strip()
                destination = m_ref.group(2).strip()
            else:
                job_ref = parts[0].strip()
                if len(non_loaded) > 1:
                    destination = non_loaded[1].strip()

    headers = [str(ws.cell(row=3, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    col_map = _map_headers(headers)
    if "vehicleReg" not in col_map:
        return None

    trucks = []
    for row in range(4, ws.max_row + 1):
        truck = _read_truck_row(ws, row, col_map, len(headers))
        if truck is None:
            break
        truck["vehicleReg"] = truck["vehicleReg"].upper()
        trucks.append(truck)

    if not trucks:
        return None

    return {
        "orderNumber": job_ref or None,
        "destinationAddress": destination or None,
        "clientName": destination or None,
        "pickupDate": loaded_date,
        "trucks": trucks,
    }


# ── Template 2: Bulk Connections standard download template ──────────────────

def _extract_bulk_template(ws) -> dict | None:
    """
    Row 1  — title containing 'BULK CONNECTIONS'
    Rows 3–9 — label | value order-info fields
    Row 11 — column headers
    Row 12+ — truck data
    """
    title = str(ws.cell(row=1, column=1).value or "")
    if "BULK CONNECTIONS" not in title.upper():
        return None

    field_map = {
        "ORDER NUMBER": "orderNumber",
        "CLIENT": "clientName", "MINE": "clientName",
        "PRODUCT": "product",
        "QUANTITY": "orderQty", "TOTAL QUANTITY": "orderQty",
        "ORIGIN": "originAddress",
        "DESTINATION": "destinationAddress",
        "PICKUP DATE": "pickupDate",
    }
    order_info: dict = {}
    for row in range(3, 10):
        label = str(ws.cell(row=row, column=1).value or "").strip().upper().rstrip(":")
        value = ws.cell(row=row, column=2).value
        if not label or not value:
            continue
        for key, fld in field_map.items():
            if key in label:
                order_info[fld] = str(value).strip()
                break

    headers = [str(ws.cell(row=11, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    col_map = _map_headers(headers)
    if "vehicleReg" not in col_map:
        return None

    trucks = []
    for row in range(12, ws.max_row + 1):
        truck = _read_truck_row(ws, row, col_map, len(headers))
        if truck is None:
            break
        truck["vehicleReg"] = truck["vehicleReg"].upper()
        trucks.append(truck)

    if not trucks:
        return None

    result = {**order_info, "trucks": trucks}
    if "pickupDate" in result:
        result["pickupDate"] = safe_date(result["pickupDate"])
    if "orderQty" in result:
        result["orderQty"] = safe_float(result["orderQty"])
    return result


# ── Template 3: Generic auto-detect ─────────────────────────────────────────

def _extract_generic_template(ws) -> dict | None:
    """
    Scan first 15 rows for a header row that contains a vehicle-reg column,
    then extract all data rows below it.
    """
    vehicle_kws = {"HORSE REG", "VEHICLE REG", "HORSE", "TRUCK REG"}
    header_row = None
    for row in range(1, min(16, ws.max_row + 1)):
        vals = {str(ws.cell(row=row, column=c).value or "").strip().upper()
                for c in range(1, ws.max_column + 1)}
        if vals & vehicle_kws:
            header_row = row
            break
    if not header_row:
        return None

    headers = [str(ws.cell(row=header_row, column=c).value or "").strip()
               for c in range(1, ws.max_column + 1)]
    col_map = _map_headers(headers)
    if "vehicleReg" not in col_map:
        return None

    trucks = []
    empty_streak = 0
    row = header_row + 1
    while row <= ws.max_row and empty_streak < 3:
        truck = _read_truck_row(ws, row, col_map, len(headers))
        if truck is None:
            empty_streak += 1
        else:
            empty_streak = 0
            truck["vehicleReg"] = truck["vehicleReg"].upper()
            trucks.append(truck)
        row += 1

    return {"trucks": trucks} if trucks else None


# ── Row-1 order info extractor (fallback for all templates) ─────────────────

def _extract_row1_order_info(ws) -> dict:
    """
    Scan row 1 of the worksheet and extract orderNumber, clientName, pickupDate.
    Expected format (all in one cell or spread across cells):
        ORDER_CODE  COMPANY NAME  LOADED DD-MM-YYYY
    Returns a dict with whichever fields were found (may be empty).
    """
    texts = _first_row_texts(ws, 1)
    combined = " ".join(texts).strip()
    result: dict = {}

    if not combined:
        return result

    date_patterns = [
        r'LOADED\s+(\d{1,2}[-/]\d{1,2}[-/]\d{4})',
        r'LOADED\s+(\d{4}[-/]\d{2}[-/]\d{2})',
        r'LOADED\s+(\d{1,2}[-/]\d{1,2}[-/]\d{2})\b',
    ]
    for pat in date_patterns:
        m = re.search(pat, combined, re.IGNORECASE)
        if m:
            d = safe_date(m.group(1))
            if d:
                result["pickupDate"] = d
            combined = re.sub(pat, '', combined, flags=re.IGNORECASE).strip()
            break

    combined = re.sub(r'\s+', ' ', combined).strip()
    if not combined:
        return result

    parts = re.split(r'\s*[–—]\s*|\s+-\s+', combined, maxsplit=1)
    if len(parts) > 1:
        result["orderNumber"] = parts[0].strip()
        result["clientName"] = parts[1].strip()
    else:
        m_ref = re.match(r'^(\S+)\s+(.+)', combined)
        if m_ref and re.search(r'\d', m_ref.group(1)):
            result["orderNumber"] = m_ref.group(1).strip()
            result["clientName"] = m_ref.group(2).strip()
        else:
            result["orderNumber"] = combined.strip()

    logger.info(f"Row-1 order info: {result}")
    return result


# ── Dispatcher ───────────────────────────────────────────────────────────────

def extract_excel_structured(file_bytes: bytes) -> dict | None:
    """
    Try each template in priority order; return the first match or None.
    None signals the caller to fall back to Gemini AI.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        for extractor, name in [
            (_extract_kfts_template,    "KFTS / Island View"),
            (_extract_bulk_template,    "Bulk Connections standard"),
            (_extract_generic_template, "Generic auto-detect"),
        ]:
            result = extractor(ws)
            if result:
                if not result.get("orderNumber") or not result.get("clientName") or not result.get("pickupDate"):
                    row1 = _extract_row1_order_info(ws)
                    for key in ("orderNumber", "clientName", "pickupDate"):
                        if not result.get(key) and row1.get(key):
                            result[key] = row1[key]
                logger.info(f"Excel matched '{name}' template — "
                            f"{len(result.get('trucks', []))} trucks, order={result.get('orderNumber')}")
                return result

        logger.info("No Excel template matched — falling back to Gemini AI")
        return None
    except Exception as e:
        logger.error(f"Structured Excel extraction error: {e}")
        return None


def excel_to_text(file_bytes: bytes) -> str:
    """Convert all sheets of an Excel file to readable text (Gemini fallback)."""
    try:
        xl = pd.ExcelFile(io.BytesIO(file_bytes))
        parts = []
        for sheet_name in xl.sheet_names:
            df = xl.parse(sheet_name, header=None, dtype=str)
            df = df.fillna("")
            parts.append(f"=== SHEET: {sheet_name} ===")
            for _, row in df.iterrows():
                line = " | ".join(str(v).strip() for v in row if str(v).strip())
                if line:
                    parts.append(line)
        return "\n".join(parts)
    except Exception as e:
        logger.error(f"Excel to text error: {e}")
        return ""
